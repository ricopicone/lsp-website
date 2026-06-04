"""Parse the treasurer's historical tuition/dues spreadsheets and match the
free-text payer names to existing :class:`accounts.User` rows.

The treasurer tracked tuition and dues in per-academic-year Excel ledgers with
no stable identifiers — just a typed name, an amount, a date, and (for tuition)
an installment label. Names are wildly inconsistent: typos (``Garet``/``Garret``
Barnwell), reordered CJK names (``Deng Xuekang`` vs the directory's
``Doreen Xuekang Deng``), partial names (``Laura Rodriguez`` vs ``Laura Rivera
Rodriguez``), and third-party payers (``Alice Wu (Payment from Joseph Scott
Rodriguez)``).

This module does two things:

* :func:`parse_tuition_ledger` / :func:`parse_dues_ledger` turn an ``.xlsx`` file
  into a list of normalized :class:`LedgerRow` records.
* :class:`NameMatcher` resolves a raw payer name to a User using only
  *high-confidence* rules (exact normalized full name, order-independent token
  match, or a unique last-name + first-initial match). Anything ambiguous returns
  ``None`` so the importer can skip-and-log it rather than guess on financial
  data.

The import command in ``payments/management/commands/import_treasurer_payments.py``
consumes both.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

# ---------------------------------------------------------------------------
# Name normalization + matching
# ---------------------------------------------------------------------------

#: Tokens that are never part of a person's name and should be dropped before
#: comparison (honorifics, generational suffixes).
_NOISE_TOKENS = {"dr", "mr", "mrs", "ms", "mx", "prof", "jr", "sr", "ii", "iii", "iv"}

#: Professional credentials that ride along on card-billing names ("Annie Rogers
#: PhD", "Nathan Lupo LMFT", "… Psych …"). Dropped before matching so the real
#: surname isn't mistaken for a credential. Deliberately excludes ambiguous
#: short tokens that are also surnames (e.g. "Ma", "Do", "MD") to avoid
#: erasing a real name.
_CREDENTIAL_TOKENS = {
    "phd", "psyd", "psya", "psyad", "lmft", "lcsw", "lcpc", "lpc", "lpcc",
    "lmhc", "lcat", "msw", "edd", "dsw", "mph", "mfa", "psych", "rn", "np",
}

#: Surname particles — kept for display but ignored for token-set matching so
#: ``Shanna Carlson de la Torre`` matches ``Shanna Carlson``.
_PARTICLES = {"de", "la", "le", "del", "della", "di", "da", "van", "von",
              "der", "den", "el", "al", "bin", "bou", "ben"}


def strip_accents(text: str) -> str:
    """Fold accented characters to ASCII (María -> Maria)."""
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def clean_raw_name(raw: str) -> str:
    """Reduce a messy ledger name to the primary payer's name.

    Strips parentheticals (``(William Chan)``, ``(Northern NM College)``),
    "payment from …" annotations, and everything after a slash (which in these
    sheets always separates a member from a third-party payer — we keep the
    member, who is listed first).
    """
    s = raw.strip()
    # Remove parenthetical asides ("(William Chan)", "(Payment from Joseph)").
    s = re.sub(r"\([^)]*\)", " ", s)
    # Drop any "payment from X" / "paid by X" annotation that wasn't parenthesized.
    s = re.sub(r"\b(payment from|paid by|on behalf of)\b.*", "", s, flags=re.I)
    # A slash separates payer alternatives — keep the first.
    if "/" in s:
        s = s.split("/", 1)[0]
    # Drop any stray unmatched bracket left over.
    s = s.replace("(", " ").replace(")", " ")
    return " ".join(s.split()).strip(" ,;")


def name_tokens(raw: str) -> list[str]:
    """Normalized, order-preserving significant tokens of a cleaned name.

    Lowercased, accent-folded, punctuation-stripped, with honorifics/suffixes
    removed. Particles are *kept* here (they matter for display) but the
    token-*set* comparison in :class:`NameMatcher` drops them.
    """
    s = strip_accents(clean_raw_name(raw)).lower()
    s = re.sub(r"[^a-z\s]", " ", s)
    toks = [
        t for t in s.split()
        if t and t not in _NOISE_TOKENS and t not in _CREDENTIAL_TOKENS
    ]
    return toks


def _core_set(tokens: list[str]) -> frozenset[str]:
    """Token set with surname particles removed, for order-independent compare."""
    return frozenset(t for t in tokens if t not in _PARTICLES)


def _alias_key(raw: str) -> str:
    """Normalize a raw ledger name to a stable alias-map key.

    Lowercase, accent-folded, punctuation→space, collapsed. Unlike
    :func:`clean_raw_name` it keeps *all* words (parentheticals included) so
    ``Wu Bing`` and ``Wu Bing (Winnie)`` stay distinct keys.
    """
    s = strip_accents(raw).lower()
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    return " ".join(s.split())


#: Human-verified raw-ledger-name → canonical-roster-name map. Each pair was
#: confirmed 1:1 against the live roster (see docs/treasurer-import.md), so this
#: stays within the "high-confidence only" matching bar — these are not fuzzy
#: guesses. Keys are :func:`_alias_key`-normalized. Covers roster
#: spelling/ordering differences (William Chan = "Chan Wai Lim", Winnie Wu =
#: "Wu Bing", Sand Jianglan = "Jiang Lan", …) and outright ledger typos
#: (Lascano→Lazcano, Hofman→Hofmann, Davison→Davidson, Bennet→Bennett, …).
ALIASES: dict[str, str] = {
    "chan wai lim william chan": "William Chan",
    "chan wai lim": "William Chan",
    "jiang lan": "Sand Jianglan",
    "cissy hong": "Cissy Hong Zhou",
    "karla martin": "Karla Martin-Gonzalez",
    "shanna carlson": "Shanna Carlson de la Torre",
    "michael tod edgerton": "Tod Edgerton",
    "liu ronan": "Ruonan Liu",
    "maria liza": "María Líza Ahearne",
    "wu bing": "Winnie Wu",
    "wu bing winnie": "Winnie Wu",
    "doreen xuekang deng": "Doreen Xuekang Deng",
    "calumn neil": "Calum Neill",
    "christopher chamberlain": "Christopher Chamberlin",
    "beatrice patsalides hofman": "Beatrice Patsalides Hofmann",
    "benjamin davison": "Benjamin Davidson",
    "hannah bennet": "Hannah Patricia Bennett",
    "danielle sawicki": "Daniela Sawicki Rivera",
    "roberto lascano": "Roberto Lazcano",
}


@dataclass
class MatchResult:
    user_id: int | None
    confidence: str  # "exact" | "tokenset" | "lastinitial" | "none"
    reason: str


class NameMatcher:
    """Resolve free-text payer names to User ids with high-confidence rules only.

    Build it from an iterable of ``(user_id, first_name, last_name,
    display_name)`` tuples (see :meth:`from_queryset`). Call :meth:`match`.
    """

    def __init__(self, people: list[tuple[int, str, str, str]]):
        # Index by exact normalized "first last" string and by core token set.
        self._by_fullname: dict[str, set[int]] = {}
        self._by_tokenset: dict[frozenset[str], set[int]] = {}
        # last-name -> {(first_initial, user_id)}
        self._by_lastname: dict[str, set[tuple[str, int]]] = {}
        # surname token -> {user_id}, for the subset rule
        self._surname_owners: dict[str, set[int]] = {}
        # user_id -> list of core token sets (one per name variant)
        self._uid_tokensets: dict[int, list[frozenset[str]]] = {}
        self._names: dict[int, str] = {}

        for uid, first, last, display in people:
            self._names[uid] = f"{first} {last}".strip() or display
            variants = {f"{first} {last}", display}
            for variant in variants:
                toks = name_tokens(variant or "")
                if not toks:
                    continue
                core_set = _core_set(toks)
                self._by_fullname.setdefault(" ".join(toks), set()).add(uid)
                self._by_tokenset.setdefault(core_set, set()).add(uid)
                self._uid_tokensets.setdefault(uid, []).append(core_set)
                core = [t for t in toks if t not in _PARTICLES]
                if core:
                    last_tok = core[-1]
                    first_initial = core[0][0]
                    self._by_lastname.setdefault(last_tok, set()).add(
                        (first_initial, uid)
                    )
                    self._surname_owners.setdefault(last_tok, set()).add(uid)

    @classmethod
    def from_queryset(cls, qs) -> NameMatcher:
        """Build from a User queryset (needs first_name, last_name, profile)."""
        people = []
        for u in qs:
            display = ""
            prof = getattr(u, "profile", None)
            if prof is not None:
                display = getattr(prof, "display_name", "") or ""
            people.append((u.pk, u.first_name or "", u.last_name or "", display))
        return cls(people)

    def display_name(self, user_id: int) -> str:
        return self._names.get(user_id, f"user#{user_id}")

    def match(self, raw: str) -> MatchResult:
        # A verified alias substitutes the canonical roster name before matching.
        canonical = ALIASES.get(_alias_key(raw))
        toks = name_tokens(canonical or raw)
        if not toks:
            return MatchResult(None, "none", "empty name")
        if canonical:
            # The alias target is human-verified; resolve it directly so a stray
            # roster collision can't silently drop a known mapping.
            exact = self._by_fullname.get(" ".join(toks))
            if exact and len(exact) == 1:
                return MatchResult(next(iter(exact)), "alias", "verified alias")
            tset = self._by_tokenset.get(_core_set(toks))
            if tset and len(tset) == 1:
                return MatchResult(next(iter(tset)), "alias", "verified alias")
            return MatchResult(
                None, "none",
                f"alias target {canonical!r} not uniquely in roster",
            )

        # 1. Exact normalized full name (after order-independent set check below
        #    we still prefer an exact string hit when unique).
        exact = self._by_fullname.get(" ".join(toks))
        if exact and len(exact) == 1:
            return MatchResult(next(iter(exact)), "exact", "exact normalized name")

        # 2. Order-independent token set (handles reordered CJK names and
        #    middle-name presence/absence collapsing to the same core set only
        #    when it's an exact set equality).
        core = _core_set(toks)
        tset = self._by_tokenset.get(core)
        if tset and len(tset) == 1:
            return MatchResult(next(iter(tset)), "tokenset", "token-set match")

        # 3. Unique last name + matching first initial. Catches partial names
        #    (Laura Rodriguez -> Laura Rivera Rodriguez) and minor first-name
        #    typos, but only when exactly one roster member shares that surname
        #    and initial.
        core_toks = [t for t in toks if t not in _PARTICLES]
        if core_toks:
            cands = self._by_lastname.get(core_toks[-1], set())
            initial = core_toks[0][0]
            hits = {uid for (fi, uid) in cands if fi == initial}
            if len(hits) == 1:
                return MatchResult(
                    next(iter(hits)), "lastinitial",
                    "unique last-name + first-initial",
                )

        # 4. Subset: the ledger name's tokens are a subset of exactly one
        #    roster member's tokens AND share that member's surname. Catches a
        #    middle name standing in for the first (``Tod Edgerton`` ->
        #    ``Michael Tod Edgerton``) and dropped given names
        #    (``Deng Xuekang`` -> ``Doreen Xuekang Deng``) without matching on a
        #    bare first name (the surname-share guard blocks that).
        if core_toks:
            surname_owners = set()
            for t in core_toks:
                surname_owners |= self._surname_owners.get(t, set())
            subset_hits = {
                uid for uid in surname_owners
                if any(core <= ts for ts in self._uid_tokensets.get(uid, []))
            }
            if len(subset_hits) == 1:
                return MatchResult(
                    next(iter(subset_hits)), "subset",
                    "unique subset-of-roster-name sharing surname",
                )

        return MatchResult(None, "none", "no high-confidence match")


# ---------------------------------------------------------------------------
# Ledger parsing
# ---------------------------------------------------------------------------


@dataclass
class LedgerRow:
    """One parsed payment row from a treasurer ledger."""

    row_index: int            # 0-based row position in the sheet (for the import tag)
    raw_name: str
    amount: Decimal
    paid_on: date | None
    installment_label: str = ""   # tuition only: "1st", "Full", ...
    method: str = ""              # dues only: "Stripe", "Paypal", ...
    note: str = ""                # any trailing free-text note
    flags: list[str] = field(default_factory=list)  # e.g. ["donation"]


def _as_decimal(value) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    s = re.sub(r"[^0-9.\-]", "", str(value))
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _as_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _iter_rows(path: Path, sheet: str | None = None):
    import openpyxl  # local import: only needed when actually parsing

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.active
    try:
        yield from enumerate(ws.iter_rows(values_only=True))
    finally:
        wb.close()


def _looks_like_header(name: str) -> bool:
    return name.strip().lower() in {"name", "names"}


def parse_tuition_ledger(path: Path) -> list[LedgerRow]:
    """Parse a tuition ledger: ``Name, Amount, Date, Installment, [paid-as note]``.

    Some files have a header row, some don't; we detect and skip it. Rows whose
    first cell isn't a name, or that have no amount, are skipped (totals, blanks).
    """
    rows: list[LedgerRow] = []
    for i, raw in _iter_rows(Path(path)):
        if not raw:
            continue
        name = raw[0]
        if not isinstance(name, str) or not name.strip():
            continue
        if _looks_like_header(name):
            continue
        amount = _as_decimal(raw[1] if len(raw) > 1 else None)
        if amount is None:
            continue
        paid_on = _as_date(raw[2] if len(raw) > 2 else None)
        installment = ""
        if len(raw) > 3 and raw[3] is not None:
            installment = str(raw[3]).strip()
        note = ""
        if len(raw) > 4 and raw[4] is not None:
            note = str(raw[4]).strip()
        rows.append(
            LedgerRow(
                row_index=i,
                raw_name=name.strip(),
                amount=amount,
                paid_on=paid_on,
                installment_label=installment,
                note=note,
            )
        )
    return rows


def parse_dues_ledger(path: Path) -> list[LedgerRow]:
    """Parse a dues ledger.

    Two shapes occur:
      * ``Name, Amount, Method, Date``  (Dues 24-25)
      * ``Name, Amount, Date, [tag]``   (Dues 25-26, where col 3 may be 'DONATION')

    We sniff which of cols 2/3 is the date vs the method. A row tagged DONATION
    (anywhere after the amount) is flagged so the importer can record it as a
    donation rather than dues. Trailing total rows (no name) are skipped.
    """
    rows: list[LedgerRow] = []
    for i, raw in _iter_rows(Path(path)):
        if not raw:
            continue
        name = raw[0]
        if not isinstance(name, str) or not name.strip():
            continue
        if _looks_like_header(name):
            continue
        amount = _as_decimal(raw[1] if len(raw) > 1 else None)
        if amount is None:
            continue

        method = ""
        paid_on = None
        flags: list[str] = []
        note = ""
        for cell in raw[2:]:
            if cell is None or cell == "":
                continue
            d = _as_date(cell)
            if d is not None and paid_on is None:
                paid_on = d
                continue
            text = str(cell).strip()
            if text.lower() in {"stripe", "paypal", "check", "cash", "offline", "venmo", "zelle"}:
                method = text
            elif "donation" in text.lower():
                flags.append("donation")
                note = (note + " " + text).strip()
            else:
                note = (note + " " + text).strip()

        rows.append(
            LedgerRow(
                row_index=i,
                raw_name=name.strip(),
                amount=amount,
                paid_on=paid_on,
                method=method,
                note=note,
                flags=flags,
            )
        )
    return rows
